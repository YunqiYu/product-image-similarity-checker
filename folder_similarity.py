from __future__ import annotations

import argparse
import base64
import csv
import json
import mimetypes
import os
import re
import time
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import Any
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from PIL import Image as PILImage

from similarity import format_duration, parse_percent


def normalize_zip_path(path_text: str) -> str:
    parts: list[str] = []
    for part in PurePosixPath(path_text).parts:
        if part == "..":
            if parts:
                parts.pop()
        elif part != ".":
            parts.append(part)
    return "/".join(parts)


DEFAULT_INPUT_DIR = Path(__file__).parent / "inputs" / "测试"
DEFAULT_OUTPUT = Path(__file__).parent / "outputs" / "测试_图片相似度结果.xlsx"
DEFAULT_RUN_LOG = Path(__file__).parent / "outputs" / "测试_图片相似度结果.log"
DEFAULT_METADATA_EXCEL = Path(__file__).parent / "inputs" / "一品红表格 - 打标核对.xlsx"
DEFAULT_METADATA_SHEET = "最终结果"
DEFAULT_PROMPT = Path(__file__).parent / "config" / "prompt.md"

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}
FALLBACK_OUTPUT_HEADERS = [
    "1级分类",
    "主题标签",
    "ASIN",
    "图片链接",
    "图片",
    "综合相似度",
    "产品链接",
    "品牌",
    "pcs",
    "标题",
    "近30天销量",
    "上架时间",
    "价格",
    "数据来源",
    "价格趋势图",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare competitor images in a folder against one target image.")
    parser.add_argument("--input-dir", default=str(DEFAULT_INPUT_DIR))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--run-log", default=str(DEFAULT_RUN_LOG))
    parser.add_argument("--model", default=None)
    parser.add_argument("--workers", type=int, default=3)
    parser.add_argument("--image-detail", default=None)
    parser.add_argument("--request-timeout", type=float, default=60.0)
    parser.add_argument("--max-retries", type=int, default=1)
    parser.add_argument("--similarity-threshold", type=float, default=80.0)
    parser.add_argument("--metadata-excel", default=str(DEFAULT_METADATA_EXCEL))
    parser.add_argument("--metadata-sheet", default=DEFAULT_METADATA_SHEET)
    parser.add_argument("--prompt", default=str(DEFAULT_PROMPT))
    return parser.parse_args()


def image_to_data_url(path: Path) -> str:
    mime_type = mimetypes.guess_type(path.name)[0] or "image/png"
    data = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime_type};base64,{data}"


def find_target_image(input_dir: Path) -> Path:
    for path in input_dir.iterdir():
        if path.is_file() and "目标" in path.stem and path.suffix.lower() in IMAGE_EXTENSIONS:
            return path
    raise FileNotFoundError(f"No target image found in {input_dir}. Expected a file name containing '目标'.")


def list_candidate_images(input_dir: Path, target_image: Path) -> list[Path]:
    images = [
        path
        for path in input_dir.iterdir()
        if path.is_file() and path != target_image and path.suffix.lower() in IMAGE_EXTENSIONS
    ]
    return sorted(images, key=lambda path: natural_key(path.stem))


def natural_key(text: str) -> list[Any]:
    parts = re.split(r"(\d+)", text)
    return [int(part) if part.isdigit() else part.lower() for part in parts]


def extract_asin_from_stem(stem: str) -> str:
    match = re.search(r"(B0[A-Z0-9]{8}|B[A-Z0-9]{9})", stem, flags=re.IGNORECASE)
    return match.group(1).upper() if match else stem


def load_manifest(input_dir: Path) -> dict[str, dict[str, str]]:
    manifest_path = input_dir / "manifest.tsv"
    if not manifest_path.exists():
        return {}
    with manifest_path.open("r", encoding="utf-8-sig", newline="") as file:
        return {
            row["filename"]: row
            for row in csv.DictReader(file, delimiter="\t")
            if row.get("filename")
        }


def load_excel_metadata(excel_path: Path, sheet_name: str) -> dict[int, dict[str, Any]]:
    if not excel_path.exists():
        return {}
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return {}
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    metadata: dict[int, dict[str, Any]] = {}
    for row_number, row_values in enumerate(rows, start=2):
        metadata[row_number] = dict(zip(headers, row_values))
    return metadata


def load_excel_headers(excel_path: Path, sheet_name: str) -> list[str]:
    if not excel_path.exists():
        return []
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return []
    worksheet = workbook[sheet_name]
    return [str(value).strip() if value is not None else "" for value in next(worksheet.iter_rows(values_only=True))]


def build_output_headers(source_headers: list[str]) -> list[str]:
    excluded_headers = {"综合相似度", "颜色", "风格", "元素", "排版"}
    headers = [
        "类别" if header == "类目路径" else header
        for header in source_headers
        if header and header not in excluded_headers
    ]
    if not headers:
        headers = [header for header in FALLBACK_OUTPUT_HEADERS if header not in excluded_headers]
    if "图片" in headers:
        insert_at = headers.index("图片") + 1
    elif "图片链接" in headers:
        insert_at = headers.index("图片链接") + 1
        headers.insert(insert_at, "图片")
        insert_at += 1
    else:
        insert_at = len(headers)
    return headers[:insert_at] + ["综合相似度"] + headers[insert_at:]


def metadata_for_candidate(
    candidate: Path,
    manifest: dict[str, dict[str, str]],
    excel_metadata: dict[int, dict[str, Any]],
) -> dict[str, Any]:
    row: dict[str, Any] = {}
    manifest_row = manifest.get(candidate.name, {})
    row_number = manifest_row.get("row")
    if row_number and row_number.isdigit():
        row["__source_row"] = int(row_number)
        row.update(excel_metadata.get(int(row_number), {}))
    if manifest_row:
        row.setdefault("ASIN", manifest_row.get("asin"))
        row.setdefault("图片链接", manifest_row.get("original_url") or manifest_row.get("used_url_size"))
    if "类别" not in row and row.get("类目路径"):
        row["类别"] = row.get("类目路径")
    row.setdefault("ASIN", extract_asin_from_stem(candidate.stem))
    row.setdefault("图片链接", candidate.name)
    return row


def extract_price_trend_images(
    excel_path: Path,
    sheet_name: str,
    cache_dir: Path,
    row_count: int,
) -> dict[int, Path]:
    if not excel_path.exists() or row_count <= 0:
        return {}
    cache_dir.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return {}
    worksheet = workbook[sheet_name]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    if "价格趋势图" not in headers:
        return {}

    with ZipFile(excel_path) as archive:
        dispimg_images = extract_dispimg_price_trend_images(excel_path, sheet_name, archive, cache_dir)
        if dispimg_images:
            return dispimg_images

        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        workbook_ns = {
            "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
        }
        workbook_rels = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels_root.findall("rel:Relationship", workbook_ns)
        }
        sheet_target = ""
        for sheet in workbook_root.findall(".//main:sheet", workbook_ns):
            if sheet.attrib.get("name") == sheet_name:
                rel_id = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
                sheet_target = workbook_rels.get(rel_id, "")
                break
        if not sheet_target:
            return {}
        sheet_path = f"xl/{sheet_target.lstrip('/')}"
        sheet_rels_path = f"{PurePosixPath(sheet_path).parent}/_rels/{PurePosixPath(sheet_path).name}.rels"
        if sheet_rels_path not in archive.namelist():
            return {}
        sheet_rels_root = ET.fromstring(archive.read(sheet_rels_path))
        drawing_target = ""
        for rel in sheet_rels_root.findall("rel:Relationship", workbook_ns):
            if rel.attrib.get("Type", "").endswith("/drawing"):
                drawing_target = rel.attrib["Target"]
                break
        if not drawing_target:
            return {}
        drawing_path = normalize_zip_path(str(PurePosixPath(sheet_path).parent / drawing_target))
        drawing_rels_path = f"{PurePosixPath(drawing_path).parent}/_rels/{PurePosixPath(drawing_path).name}.rels"
        if drawing_path not in archive.namelist() or drawing_rels_path not in archive.namelist():
            return {}

        drawing_rels_root = ET.fromstring(archive.read(drawing_rels_path))
        drawing_rels = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in drawing_rels_root.findall("rel:Relationship", workbook_ns)
        }
        drawing_root = ET.fromstring(archive.read(drawing_path))
        drawing_ns = {
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        }
        trend_images: dict[int, Path] = {}
        blips = drawing_root.findall(".//a:blip", drawing_ns)
        for offset, blip in enumerate(blips[:row_count], start=2):
            rel_id = blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
            target = drawing_rels.get(rel_id or "")
            if not target:
                continue
            media_path = normalize_zip_path(str(PurePosixPath(drawing_path).parent / target))
            if media_path not in archive.namelist():
                continue
            with PILImage.open(BytesIO(archive.read(media_path))) as source:
                image_path = cache_dir / f"trend_row_{offset}.png"
                source.convert("RGB").save(image_path, "PNG")
            trend_images[offset] = image_path
    return trend_images


def extract_dispimg_price_trend_images(
    excel_path: Path,
    sheet_name: str,
    archive: ZipFile,
    cache_dir: Path,
) -> dict[int, Path]:
    if "xl/cellimages.xml" not in archive.namelist() or "xl/_rels/cellimages.xml.rels" not in archive.namelist():
        return {}

    workbook = load_workbook(excel_path, read_only=True, data_only=False)
    if sheet_name not in workbook.sheetnames:
        return {}
    worksheet = workbook[sheet_name]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    if "价格趋势图" not in headers:
        return {}
    trend_col = headers.index("价格趋势图") + 1

    package_ns = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}
    rels_root = ET.fromstring(archive.read("xl/_rels/cellimages.xml.rels"))
    rels = {
        rel.attrib["Id"]: normalize_zip_path(f"xl/{rel.attrib['Target']}")
        for rel in rels_root.findall("rel:Relationship", package_ns)
    }

    cell_ns = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    }
    cellimages_root = ET.fromstring(archive.read("xl/cellimages.xml"))
    image_id_to_media: dict[str, str] = {}
    for pic in cellimages_root.findall(".//xdr:pic", cell_ns):
        name_node = pic.find(".//xdr:cNvPr", cell_ns)
        blip = pic.find(".//a:blip", cell_ns)
        if name_node is None or blip is None:
            continue
        image_id = name_node.attrib.get("name")
        rel_id = blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
        media_path = rels.get(rel_id or "")
        if image_id and media_path:
            image_id_to_media[image_id] = media_path

    trend_images: dict[int, Path] = {}
    pattern = re.compile(r'DISPIMG\("([^"]+)"')
    for row in range(2, worksheet.max_row + 1):
        value = str(worksheet.cell(row=row, column=trend_col).value or "")
        match = pattern.search(value)
        if not match:
            continue
        media_path = image_id_to_media.get(match.group(1))
        if not media_path or media_path not in archive.namelist():
            continue
        with PILImage.open(BytesIO(archive.read(media_path))) as source:
            image_path = cache_dir / f"trend_row_{row}.png"
            source.convert("RGB").save(image_path, "PNG")
        trend_images[row] = image_path
    return trend_images


def default_compare_prompt_template() -> str:
    return """Compare the target product image with the competitor product image.
Return only valid JSON with these keys:
{
  "color相似度": 0,
  "style相似度": 0,
  "elements相似度": 0,
  "format相似度": 0,
  "相似理由": ""
}
Similarity threshold: {threshold}%.
"""


def load_prompt_template(prompt_path: Path) -> str:
    if prompt_path.exists():
        return prompt_path.read_text(encoding="utf-8")
    return default_compare_prompt_template()


def build_compare_prompt(threshold: float, prompt_template: str) -> str:
    return prompt_template.replace("{threshold}", f"{threshold:g}")


def normalize_similarity_payload(data: dict[str, Any], threshold: float) -> dict[str, str]:
    color = parse_percent(data.get("color相似度", data.get("颜色相似度", data.get("color_similarity"))))
    style = parse_percent(data.get("style相似度", data.get("风格相似度", data.get("style_similarity"))))
    elements = parse_percent(data.get("elements相似度", data.get("元素相似度", data.get("elements_similarity"))))
    fmt = parse_percent(data.get("format相似度", data.get("版式相似度", data.get("format_similarity"))))
    overall = elements * 0.20 + style * 0.30 + color * 0.20 + fmt * 0.30
    return {
        "综合相似度": f"{overall:.1f}%",
        "相似理由": str(data.get("相似理由", data.get("reason", ""))).strip(),
        "color相似度": f"{round(color)}%",
        "style相似度": f"{round(style)}%",
        "elements相似度": f"{round(elements)}%",
        "format相似度": f"{round(fmt)}%",
        "是否相似": "是" if overall >= threshold else "否",
    }


def compare_pair(
    client: Any,
    model: str,
    target_url: str,
    candidate_url: str,
    image_detail: str,
    threshold: float,
    prompt_template: str,
) -> dict[str, str]:
    from tagger import extract_json

    response = client.chat.completions.create(
        model=model,
        response_format={"type": "json_object"},
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": build_compare_prompt(threshold, prompt_template)},
                    {"type": "image_url", "image_url": {"url": target_url, "detail": image_detail}},
                    {"type": "image_url", "image_url": {"url": candidate_url, "detail": image_detail}},
                ],
            }
        ],
    )
    return normalize_similarity_payload(extract_json(response.choices[0].message.content or ""), threshold)


def percent_sort_value(value: Any) -> float:
    try:
        return parse_percent(value)
    except Exception:
        return -1.0


def make_excel_thumbnail(image_path: Path, thumbnail_dir: Path, size: int = 500) -> Path:
    thumbnail_dir.mkdir(parents=True, exist_ok=True)
    thumbnail_path = thumbnail_dir / f"{image_path.stem}.png"
    with PILImage.open(image_path) as source:
        source.thumbnail((size, size))
        canvas = PILImage.new("RGB", (size, size), "white")
        x = (size - source.width) // 2
        y = (size - source.height) // 2
        canvas.paste(source.convert("RGB"), (x, y))
        canvas.save(thumbnail_path, "PNG")
    return thumbnail_path


def add_image(ws: Any, image_path: Path, cell: str, size: int = 220) -> None:
    """Embed the downloaded original, while keeping its worksheet display compact."""
    image = ExcelImage(str(image_path))
    image.width = size
    image.height = size
    ws.add_image(image, cell)


def add_resized_image(ws: Any, image_path: Path, cell: str, width: int, height: int) -> None:
    image = ExcelImage(str(image_path))
    image.width = width
    image.height = height
    ws.add_image(image, cell)


def process_candidate(
    candidate: Path,
    client: Any,
    model: str,
    target_url: str,
    image_detail: str,
    threshold: float,
    prompt_template: str,
) -> dict[str, Any]:
    started = time.perf_counter()
    try:
        similarities = compare_pair(
            client,
            model,
            target_url,
            image_to_data_url(candidate),
            image_detail,
            threshold,
            prompt_template,
        )
        error = ""
    except Exception as exc:
        similarities = {
            "综合相似度": "",
            "相似理由": "",
            "color相似度": "",
            "style相似度": "",
            "elements相似度": "",
            "format相似度": "",
            "是否相似": "",
        }
        error = str(exc)
    return {
        "path": candidate,
        "similarities": similarities,
        "error": error,
        "elapsed": time.perf_counter() - started,
    }


def main() -> int:
    from dotenv import load_dotenv
    from openai import OpenAI

    args = parse_args()
    load_dotenv(Path(__file__).parent / ".env")

    input_dir = Path(args.input_dir)
    output_path = Path(args.output)
    run_log_path = Path(args.run_log)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    run_log_path.parent.mkdir(parents=True, exist_ok=True)
    run_log_path.write_text("", encoding="utf-8")

    target_image = find_target_image(input_dir)
    candidates = list_candidate_images(input_dir, target_image)
    if not candidates:
        raise RuntimeError(f"No candidate images found in {input_dir}")
    manifest = load_manifest(input_dir)
    excel_metadata = load_excel_metadata(Path(args.metadata_excel), args.metadata_sheet)
    source_headers = load_excel_headers(Path(args.metadata_excel), args.metadata_sheet)
    output_headers = build_output_headers(source_headers)

    api_key = os.getenv("OPENAI_API_KEY") or os.getenv("TEXT_API_KEY")
    if not api_key:
        raise RuntimeError("Missing OPENAI_API_KEY or TEXT_API_KEY")
    base_url = os.getenv("OPENAI_BASE_URL") or os.getenv("LLM_BASE_URL")
    model = args.model or os.getenv("OPENAI_MODEL") or "gpt-5.5"
    image_detail = args.image_detail or os.getenv("OPENAI_IMAGE_MODE") or "high"
    client = OpenAI(api_key=api_key, base_url=base_url, timeout=args.request_timeout, max_retries=args.max_retries)
    prompt_template = load_prompt_template(Path(args.prompt))

    target_url = image_to_data_url(target_image)
    run_started = time.perf_counter()

    results: dict[Path, dict[str, Any]] = {}
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {
            executor.submit(
                process_candidate,
                candidate,
                client,
                model,
                target_url,
                image_detail,
                args.similarity_threshold,
                prompt_template,
            ): candidate
            for candidate in candidates
        }
        done = 0
        for future in as_completed(futures):
            candidate = futures[future]
            result = future.result()
            results[candidate] = result
            done += 1
            with run_log_path.open("a", encoding="utf-8") as log:
                log.write(
                    f"{candidate.name}\t{json.dumps(result['similarities'], ensure_ascii=False)}"
                    f"\t运行时间: {format_duration(result['elapsed'])}"
                    + (f"\t错误: {result['error']}" if result["error"] else "")
                    + "\n"
                )
            print(f"[{done}/{len(candidates)}] {candidate.name} {json.dumps(result['similarities'], ensure_ascii=False)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "相似度结果"
    for col, header in enumerate(output_headers, start=1):
        ws.cell(row=1, column=col).value = header
    width_by_header = {
        "1级分类": 14,
        "主题标签": 18,
        "基本主题": 18,
        "ASIN": 18,
        "图片链接": 46,
        "图片": 32,
        "综合相似度": 14,
        "产品链接": 42,
        "品牌": 18,
        "pcs": 10,
        "标题": 46,
        "近30天销量": 14,
        "上架时间": 16,
        "价格": 12,
        "数据来源": 16,
        "类别": 16,
        "价格趋势图": 52,
    }
    for col_index, header in enumerate(output_headers, start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = width_by_header.get(header, 16)

    sorted_candidates = sorted(
        candidates,
        key=lambda candidate: percent_sort_value(results[candidate]["similarities"].get("综合相似度")),
        reverse=True,
    )
    thumbnail_dir = output_path.parent / "_excel_image_cache" / output_path.stem
    trend_images = extract_price_trend_images(
        Path(args.metadata_excel),
        args.metadata_sheet,
        thumbnail_dir / "price_trends",
        len(excel_metadata),
    )
    header_to_column = {header: index for index, header in enumerate(output_headers, start=1)}
    image_column = header_to_column.get("图片")
    overall_column = header_to_column["综合相似度"]
    trend_column = header_to_column.get("价格趋势图")

    for row_index, candidate in enumerate(sorted_candidates, start=2):
        result = results[candidate]
        similarities = result["similarities"]
        source_row = metadata_for_candidate(candidate, manifest, excel_metadata)
        for col_index, header in enumerate(output_headers, start=1):
            if header in {"图片", "综合相似度"}:
                continue
            ws.cell(row=row_index, column=col_index).value = source_row.get(header, "")
        if "ASIN" in header_to_column:
            ws.cell(row=row_index, column=header_to_column["ASIN"]).value = (
                source_row.get("ASIN") or extract_asin_from_stem(candidate.stem)
            )
        ws.row_dimensions[row_index].height = 180
        if image_column:
            ws.cell(row=row_index, column=image_column).value = candidate.name
            ws.cell(row=row_index, column=image_column).alignment = Alignment(horizontal="center", vertical="center")
            add_image(ws, candidate, f"{get_column_letter(image_column)}{row_index}")
        source_row_number = source_row.get("__source_row")
        if trend_column and source_row_number in trend_images:
            ws.cell(row=row_index, column=trend_column).value = ""
            add_resized_image(
                ws,
                trend_images[source_row_number],
                f"{get_column_letter(trend_column)}{row_index}",
                width=360,
                height=180,
            )
        ws.cell(row=row_index, column=overall_column).value = similarities["综合相似度"]

    wb.save(output_path)
    with run_log_path.open("a", encoding="utf-8") as log:
        log.write(f"本次运行总时间: {format_duration(time.perf_counter() - run_started)}\n")
    print(f"Target image: {target_image}")
    print(f"Output: {output_path}")
    print(f"Run log: {run_log_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
