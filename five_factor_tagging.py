"""五要素打标入口：读取相似度中间 Excel，为保留的商品图片补齐五个标签列。"""

from __future__ import annotations

import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from product_image_pipeline.five_factor_tagger import FIVE_FACTOR_PROMPT, tag_image
from product_image_pipeline.settings import ModelSettings
from product_image_pipeline.vision_client import create_client, image_to_data_url


LABEL_COLUMNS = {
    "颜色": "color",
    "风格": "style",
    "排版": "format",
    "元素": "elements",
    "艺术字": "wordArt",
}


def parse_args() -> argparse.Namespace:
    """解析中间 Excel、图片目录和最终 Excel 的本地调试参数。"""
    parser = argparse.ArgumentParser(description="为相似度中间 Excel 补齐五要素标签。")
    parser.add_argument("--input-excel", required=True)
    parser.add_argument("--input-dir", required=True, help="包含中间 Excel 中图片文件名的本地图片目录。")
    parser.add_argument("--output", required=True)
    parser.add_argument("--sheet", default="相似度结果")
    parser.add_argument("--image-column", default="图片")
    parser.add_argument("--model", default=None)
    parser.add_argument("--prompt", default=str(Path(__file__).parent / "config" / "five_factor_tagging_prompt.md"))
    parser.add_argument("--workers", type=int, default=3)
    parser.add_argument("--request-timeout", type=float, default=180.0)
    parser.add_argument("--max-retries", type=int, default=1)
    return parser.parse_args()


def load_prompt(path: Path) -> str:
    """优先读取可维护的 Prompt 文件，缺失时使用内置默认内容。"""
    return path.read_text(encoding="utf-8") if path.exists() else FIVE_FACTOR_PROMPT


def ensure_label_columns(worksheet: Any) -> dict[str, int]:
    """在不改变既有列顺序的前提下，将五要素列追加到表格末尾。"""
    headers = {str(cell.value).strip(): cell.column for cell in worksheet[1] if cell.value is not None}
    for header in LABEL_COLUMNS:
        if header not in headers:
            column = worksheet.max_column + 1
            worksheet.cell(row=1, column=column).value = header
            headers[header] = column
    return {header: headers[header] for header in LABEL_COLUMNS}


def format_tag_value(value: Any) -> str:
    """将列表型标签转换为 Excel 中便于业务阅读的顿号分隔文本。"""
    if isinstance(value, list):
        return "、".join(str(item) for item in value if str(item).strip())
    return str(value or "")


def main() -> int:
    """执行 Top 300 中间结果的五要素标签补齐并输出最终 Excel。"""
    from dotenv import load_dotenv

    args = parse_args()
    load_dotenv(Path(__file__).parent / ".env")
    input_excel = Path(args.input_excel)
    input_dir = Path(args.input_dir)
    output_path = Path(args.output)
    if not input_excel.exists():
        raise FileNotFoundError(f"找不到中间 Excel：{input_excel}")
    if not input_dir.exists():
        raise FileNotFoundError(f"找不到图片目录：{input_dir}")

    workbook = load_workbook(input_excel)
    if args.sheet not in workbook.sheetnames:
        raise ValueError(f"找不到工作表：{args.sheet}")
    worksheet = workbook[args.sheet]
    headers = {str(cell.value).strip(): cell.column for cell in worksheet[1] if cell.value is not None}
    if args.image_column not in headers:
        raise ValueError(f"找不到图片列：{args.image_column}")
    label_columns = ensure_label_columns(worksheet)
    image_column = headers[args.image_column]

    tasks: list[tuple[int, Path]] = []
    for row in range(2, worksheet.max_row + 1):
        image_name = str(worksheet.cell(row=row, column=image_column).value or "").strip()
        image_path = input_dir / image_name
        if image_name and image_path.exists():
            tasks.append((row, image_path))

    settings = ModelSettings.from_environment()
    client = create_client(settings, timeout=args.request_timeout, max_retries=args.max_retries)
    model = args.model or settings.tagging_model
    prompt = load_prompt(Path(args.prompt))
    results: dict[int, dict[str, Any]] = {}
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {
            executor.submit(tag_image, client, model, image_to_data_url(image_path), settings.image_detail, prompt): row
            for row, image_path in tasks
        }
        for future in as_completed(futures):
            row = futures[future]
            results[row] = future.result()

    for row, tags in results.items():
        for header, key in LABEL_COLUMNS.items():
            worksheet.cell(row=row, column=label_columns[header]).value = format_tag_value(tags[key])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"五要素打标完成: {len(results)}/{len(tasks)} 行，输出: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
