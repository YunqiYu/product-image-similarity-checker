"""五要素打标管线：读取相似度中间 Excel，为保留图片补齐五个标签列。"""

from __future__ import annotations

import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from product_image_pipeline.five_factor_tagger import FIVE_FACTOR_PROMPT, tag_image
from product_image_pipeline.settings import ModelSettings
from product_image_pipeline.similarity_ranker import add_image, add_resized_image, extract_price_trend_images, load_excel_context
from product_image_pipeline.vision_client import create_client, image_to_data_url


PROJECT_ROOT = Path(__file__).resolve().parents[1]


LABEL_COLUMNS = {
    "风格": "style",
    "颜色": "color",
    "元素": "elements",
    "排版": "format",
    "艺术字": "wordArt",
}

# 最终交付表使用固定业务字段顺序；源表暂缺的字段保持为空，等待后续系统补充。
FINAL_OUTPUT_HEADERS = [
    "1级分类",
    "主题标签",
    "基本主题",
    "ASIN",
    "图片链接",
    "图片",
    "综合相似度",
    "产品链接",
    "品牌",
    "pcs",
    "标题",
    "月份1",
    "月份2",
    "月份3",
    "总销量",
    "销量趋势图",
    "上架时间",
    "价格",
    "数据来源",
    "一起购买1",
    "一起购买2",
    "类别",
    "价格趋势图",
    "风格",
    "颜色",
    "元素",
    "排版",
    "艺术字",
]

def parse_args() -> argparse.Namespace:
    """解析中间 JSON、图片目录和最终 Excel 的本地调试参数。"""
    parser = argparse.ArgumentParser(description="为相似度中间 JSON 补齐五要素标签。")
    parser.add_argument("--input-json", required=True)
    parser.add_argument("--input-dir", required=True, help="包含中间 JSON 中图片文件名的本地图片目录。")
    parser.add_argument("--source-excel", required=True, help="用于回填完整业务字段和趋势图的原始 Excel。")
    parser.add_argument("--source-sheet", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--run-log", default=None, help="逐行打标状态日志；默认与输出 Excel 同名。")
    parser.add_argument("--model", default=None)
    parser.add_argument("--prompt", default=str(PROJECT_ROOT / "config" / "five_factor_tagging_prompt.md"))
    parser.add_argument("--workers", type=int, default=3)
    parser.add_argument("--request-timeout", type=float, default=180.0)
    parser.add_argument("--max-retries", type=int, default=1)
    parser.add_argument("--tagging-attempts", type=int, default=3, help="单张图片因接口或标签校验失败时的最大尝试次数。")
    return parser.parse_args()


def load_prompt(path: Path) -> str:
    """优先读取可维护的 Prompt 文件，缺失时使用内置默认内容。"""
    return path.read_text(encoding="utf-8") if path.exists() else FIVE_FACTOR_PROMPT


def load_intermediate_items(input_path: Path) -> list[dict[str, Any]]:
    """读取阶段一 JSON，并校验阶段二回填所需的最小字段。"""
    try:
        payload = json.loads(input_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as error:
        raise ValueError(f"中间 JSON 格式不合法：{input_path}") from error
    items = payload.get("items") if isinstance(payload, dict) else None
    if not isinstance(items, list):
        raise ValueError("中间 JSON 缺少 items 数组。")

    normalized: list[dict[str, Any]] = []
    for index, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"中间 JSON 第 {index} 条不是对象。")
        source_row = item.get("source_row")
        image_name = item.get("image_name")
        if not isinstance(source_row, int) or not isinstance(image_name, str) or not image_name.strip():
            raise ValueError(f"中间 JSON 第 {index} 条缺少 source_row 或 image_name。")
        normalized.append({
            "source_row": source_row,
            "asin": str(item.get("asin") or "").strip().upper(),
            "image_name": image_name.strip(),
            "overall": item.get("overall_similarity", ""),
            "tags": {},
        })
    return normalized


def ensure_label_columns(worksheet: Any) -> dict[str, int]:
    """在不改变既有列顺序的前提下，将五要素列追加到表格末尾。"""
    headers = {str(cell.value).strip(): cell.column for cell in worksheet[1] if cell.value is not None}
    for header in LABEL_COLUMNS:
        if header not in headers:
            column = worksheet.max_column + 1
            worksheet.cell(row=1, column=column).value = header
            headers[header] = column
    return {header: headers[header] for header in LABEL_COLUMNS}


def format_tag_value(value: Any) -> str:
    """将列表型标签转换为 Excel 中便于业务阅读的顿号分隔文本。"""
    if isinstance(value, list):
        return "、".join(str(item) for item in value if str(item).strip())
    return str(value or "")


def build_final_headers(_: list[str]) -> list[str]:
    """返回业务确认的固定最终表头，避免源 Excel 字段顺序影响交付格式。"""
    return list(FINAL_OUTPUT_HEADERS)


def write_final_excel(
    output_path: Path,
    intermediate_rows: list[dict[str, Any]],
    source_headers: list[str],
    source_metadata: dict[int, dict[str, Any]],
    trend_images: dict[int, Path],
    input_dir: Path,
) -> None:
    """回填完整业务字段，并在最终表中写入高清商品图、趋势图和五要素标签。"""
    headers = build_final_headers(source_headers)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "最终结果"
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column).value = header
    widths = {"ASIN": 18, "图片链接": 46, "图片": 32, "综合相似度": 14, "产品链接": 42, "标题": 46, "价格趋势图": 52}
    for column, header in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = widths.get(header, 16)
    columns = {header: index for index, header in enumerate(headers, start=1)}

    for row_number, item in enumerate(intermediate_rows, start=2):
        source_row = source_metadata.get(item["source_row"], {})
        for column, header in enumerate(headers, start=1):
            if header not in {"图片", "综合相似度", "价格趋势图", *LABEL_COLUMNS}:
                worksheet.cell(row=row_number, column=column).value = source_row.get(header, "")
        worksheet.cell(row=row_number, column=columns["综合相似度"]).value = item["overall"]
        for header, key in LABEL_COLUMNS.items():
            worksheet.cell(row=row_number, column=columns[header]).value = format_tag_value(item["tags"].get(key, ""))

        worksheet.row_dimensions[row_number].height = 180
        image_path = input_dir / item["image_name"]
        worksheet.cell(row=row_number, column=columns["图片"]).value = item["image_name"]
        worksheet.cell(row=row_number, column=columns["图片"]).alignment = Alignment(horizontal="center", vertical="center")
        if image_path.exists():
            add_image(worksheet, image_path, f"{get_column_letter(columns['图片'])}{row_number}")
        if "价格趋势图" in columns and item["source_row"] in trend_images:
            add_resized_image(
                worksheet,
                trend_images[item["source_row"]],
                f"{get_column_letter(columns['价格趋势图'])}{row_number}",
                width=360,
                height=180,
            )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> int:
    """执行 Top 300 中间结果的五要素标签补齐并输出最终 Excel。"""
    from dotenv import load_dotenv

    args = parse_args()
    load_dotenv(PROJECT_ROOT / ".env")
    input_json = Path(args.input_json)
    input_dir = Path(args.input_dir)
    source_excel = Path(args.source_excel)
    output_path = Path(args.output)
    run_log_path = Path(args.run_log) if args.run_log else output_path.with_suffix(".log")
    if not input_json.exists():
        raise FileNotFoundError(f"找不到中间 JSON：{input_json}")
    if not input_dir.exists():
        raise FileNotFoundError(f"找不到图片目录：{input_dir}")
    if not source_excel.exists():
        raise FileNotFoundError(f"找不到源 Excel：{source_excel}")

    intermediate_rows = load_intermediate_items(input_json)
    source_headers, source_metadata = load_excel_context(source_excel, args.source_sheet)
    tasks: list[tuple[int, Path]] = []
    for index, item in enumerate(intermediate_rows):
        image_path = input_dir / item["image_name"]
        if image_path.exists():
            tasks.append((index, image_path))

    settings = ModelSettings.from_environment()
    client = create_client(settings, timeout=args.request_timeout, max_retries=args.max_retries)
    model = args.model or settings.tagging_model
    prompt = load_prompt(Path(args.prompt))
    results: dict[int, dict[str, Any]] = {}
    failures: dict[int, str] = {}
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {
            executor.submit(
                tag_image,
                client,
                model,
                image_to_data_url(image_path),
                settings.image_detail,
                prompt,
                attempts=args.tagging_attempts,
            ): row
            for row, image_path in tasks
        }
        for future in as_completed(futures):
            row = futures[future]
            try:
                results[row] = future.result()
            except Exception as exc:
                # 单张图片失败不应阻断整批 Top 结果的最终 Excel 输出。
                failures[row] = str(exc)

    for index, item in enumerate(intermediate_rows):
        item["tags"] = results.get(index, {})

    source_rows = {item["source_row"] for item in intermediate_rows}
    trend_images = extract_price_trend_images(
        source_excel,
        args.source_sheet,
        output_path.parent / "_excel_image_cache" / output_path.stem / "price_trends",
        source_rows,
    )
    write_final_excel(output_path, intermediate_rows, source_headers, source_metadata, trend_images, input_dir)
    run_log_path.parent.mkdir(parents=True, exist_ok=True)
    with run_log_path.open("w", encoding="utf-8") as log:
        log.write(f"五要素打标成功: {len(results)} 行\n")
        log.write(f"五要素打标失败: {len(failures)} 行\n")
        for row, error in sorted(failures.items()):
            log.write(f"row{row}\t错误: {error}\n")
    print(f"五要素打标完成: 成功 {len(results)} 行，失败 {len(failures)} 行，输出: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
