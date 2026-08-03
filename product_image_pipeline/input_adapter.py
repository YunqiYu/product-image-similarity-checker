"""数据准备：将上传 Excel 中的图片链接抽样、下载并转换为排序模块可读取的本地图片目录。"""

from __future__ import annotations

import csv
import random
import re
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import load_workbook


@dataclass(frozen=True)
class PipelineInput:
    """描述一次任务上传后的文件位置和 Excel 字段约定。"""

    target_image_path: Path
    source_excel_path: Path
    sheet_name: str
    image_url_column: str = "图片链接"


@dataclass
class CandidateRecord:
    """保存候选图片与源 Excel 行之间的可追溯关系。"""

    source_row: int
    image_url: str
    filename: str = ""
    downloaded_url: str = ""
    error: str = ""


def read_candidate_records(source_excel_path: Path, sheet_name: str, image_url_column: str) -> list[CandidateRecord]:
    """读取上传 Excel 的图片链接列，不下载也不修改原始文件。"""
    workbook = load_workbook(source_excel_path, read_only=True, data_only=False)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"找不到工作表：{sheet_name}")
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    if image_url_column not in headers:
        raise ValueError(f"找不到图片链接列：{image_url_column}")
    image_column = headers.index(image_url_column)
    return [
        CandidateRecord(source_row=row_number, image_url=str(row[image_column]).strip())
        for row_number, row in enumerate(rows, start=2)
        if image_column < len(row) and str(row[image_column] or "").strip()
    ]


def sample_candidate_records(
    records: list[CandidateRecord],
    max_candidates: int,
    random_seed: int | None = None,
    selection_mode: str = "random",
) -> list[CandidateRecord]:
    """在下载前按随机或源表顺序抽样，控制网络、磁盘和模型调用成本。"""
    if max_candidates < 1:
        raise ValueError("最大候选数量必须至少为 1。")
    if len(records) <= max_candidates:
        return records
    if selection_mode == "first":
        return records[:max_candidates]
    if selection_mode != "random":
        raise ValueError("selection_mode 只能是 random 或 first。")
    generator = random.Random(random_seed) if random_seed is not None else random.SystemRandom()
    return generator.sample(records, max_candidates)


def upgrade_amazon_image_url(url: str) -> str:
    """尽可能将亚马逊 CDN 缩略图 URL 升级为高分辨率原图 URL。"""
    if "images-na.ssl-images-amazon.com" not in url and "m.media-amazon.com" not in url:
        return url
    return re.sub(r"\._[^/]+?_\.(?:jpe?g|png|webp)", "._AC_SL2000_.jpg", url, flags=re.IGNORECASE)


def filename_for_record(record: CandidateRecord) -> str:
    """根据源行号生成稳定文件名，避免不同 URL 的文件名发生冲突。"""
    suffix = Path(urlparse(record.image_url).path).suffix.lower()
    return f"row_{record.source_row}{suffix if suffix in {'.jpg', '.jpeg', '.png', '.webp'} else '.jpg'}"


def download_record(record: CandidateRecord, images_dir: Path, timeout: float) -> CandidateRecord:
    """下载一张候选原图；失败信息保留在记录中而不抛出到整批任务。"""
    record.downloaded_url = upgrade_amazon_image_url(record.image_url)
    record.filename = filename_for_record(record)
    try:
        request = urllib.request.Request(record.downloaded_url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(request, timeout=timeout) as response:
            content = response.read()
        if not content:
            raise ValueError("下载内容为空")
        (images_dir / record.filename).write_bytes(content)
    except Exception as exc:
        record.error = str(exc)
    return record


def write_manifest(records: list[CandidateRecord], manifest_path: Path) -> None:
    """写入下载映射表，供排序输出回填源 Excel 字段和趋势图使用。"""
    with manifest_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=["row", "filename", "original_url", "used_url_size", "error"], delimiter="\t")
        writer.writeheader()
        for record in records:
            writer.writerow({
                "row": record.source_row,
                "filename": record.filename,
                "original_url": record.image_url,
                "used_url_size": record.downloaded_url,
                "error": record.error,
            })


def prepare_uploaded_excel(
    pipeline_input: PipelineInput,
    images_dir: Path,
    *,
    max_candidates: int = 1000,
    random_seed: int | None = None,
    selection_mode: str = "random",
    download_workers: int = 8,
    timeout: float = 60.0,
) -> list[CandidateRecord]:
    """完成上传 Excel 的抽样、原图下载和 manifest 生成，返回成功下载的候选记录。"""
    if not pipeline_input.target_image_path.exists():
        raise FileNotFoundError(f"找不到目标图片：{pipeline_input.target_image_path}")
    if not pipeline_input.source_excel_path.exists():
        raise FileNotFoundError(f"找不到源 Excel：{pipeline_input.source_excel_path}")
    images_dir.mkdir(parents=True, exist_ok=True)
    records = sample_candidate_records(
        read_candidate_records(pipeline_input.source_excel_path, pipeline_input.sheet_name, pipeline_input.image_url_column),
        max_candidates,
        random_seed,
        selection_mode,
    )
    with ThreadPoolExecutor(max_workers=max(1, download_workers)) as executor:
        futures = [executor.submit(download_record, record, images_dir, timeout) for record in records]
        downloaded = [future.result() for future in as_completed(futures)]
    write_manifest(downloaded, images_dir / "manifest.tsv")
    return [record for record in downloaded if not record.error]
