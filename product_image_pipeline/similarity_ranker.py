"""相似度排序核心：完成抽样、模型比较和内部 Top 结果 JSON 生成。"""

from __future__ import annotations

import argparse
import csv
import json
import os
import random
import re
import time
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import Any
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage

from .models import SimilarityWeights
from .settings import ModelSettings
from .vision_client import create_client, extract_json, image_to_data_url

def normalize_zip_path(path_text: str) -> str:
    parts: list[str] = []
    for part in PurePosixPath(path_text).parts:
        if part == "..":
            if parts:
                parts.pop()
        elif part != ".":
            parts.append(part)
    return "/".join(parts)


def parse_percent(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value or "").strip().replace("%", ""))
    except ValueError:
        return 0.0


def format_duration(seconds: float) -> str:
    total_seconds = int(round(seconds))
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT_DIR = PROJECT_ROOT / "inputs" / "测试"
DEFAULT_OUTPUT = PROJECT_ROOT / "outputs" / "测试_相似度排序.json"
DEFAULT_RUN_LOG = PROJECT_ROOT / "outputs" / "测试_图片相似度结果.log"
DEFAULT_METADATA_EXCEL = PROJECT_ROOT / "inputs" / "一品红表格 - 打标核对.xlsx"
DEFAULT_METADATA_SHEET = "最终结果"
DEFAULT_PROMPT = PROJECT_ROOT / "config" / "similarity_prompt.md"

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare competitor images in a folder against one target image.")
    parser.add_argument("--input-dir", default=str(DEFAULT_INPUT_DIR))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--run-log", default=str(DEFAULT_RUN_LOG))
    parser.add_argument("--model", default=None)
    parser.add_argument("--workers", type=int, default=3)
    parser.add_argument("--image-detail", default=None)
    parser.add_argument("--request-timeout", type=float, default=60.0)
    parser.add_argument("--max-retries", type=int, default=1)
    parser.add_argument("--similarity-threshold", type=float, default=80.0)
    # 前端可传 JSON 权重，例如：{"color":20,"style":30,"elements":20,"format":30}。
    parser.add_argument("--weights", default=None, help="JSON similarity weights supplied by the calling system.")
    # Excel输入量过大时随机抽样，控制模型调用数量和 Token 成本。（默认1000抽样1000条）
    parser.add_argument("--max-candidates", type=int, default=1000, help="Maximum random candidate images to send to the model.")
    # 限制阶段二最多处理300条，控制高精度打标的调用量和最终表规模。
    parser.add_argument("--max-results", type=int, default=300, help="Maximum ranked items to keep in the internal JSON.")
    parser.add_argument("--metadata-excel", default=str(DEFAULT_METADATA_EXCEL))
    parser.add_argument("--metadata-sheet", default=DEFAULT_METADATA_SHEET)
    parser.add_argument("--prompt", default=str(DEFAULT_PROMPT))
    return parser.parse_args()


def find_target_image(input_dir: Path) -> Path:
    for path in input_dir.iterdir():
        if path.is_file() and "目标" in path.stem and path.suffix.lower() in IMAGE_EXTENSIONS:
            return path
    raise FileNotFoundError(f"No target image found in {input_dir}. Expected a file name containing '目标'.")


def list_candidate_images(input_dir: Path, target_image: Path) -> list[Path]:
    images = [
        path
        for path in input_dir.iterdir()
        if path.is_file() and path != target_image and path.suffix.lower() in IMAGE_EXTENSIONS
    ]
    return sorted(images, key=lambda path: natural_key(path.stem))


def sample_candidates(candidates: list[Path], max_candidates: int) -> list[Path]:
    if len(candidates) <= max_candidates:
        return candidates
    # 别偷懒用同一个随机种子（random seed）或者缓存的样本跑测试
    return random.sample(candidates, max_candidates)


def natural_key(text: str) -> list[Any]:
    parts = re.split(r"(\d+)", text)
    return [int(part) if part.isdigit() else part.lower() for part in parts]


def extract_asin_from_stem(stem: str) -> str:
    match = re.search(r"(B0[A-Z0-9]{8}|B[A-Z0-9]{9})", stem, flags=re.IGNORECASE)
    return match.group(1).upper() if match else stem


def load_manifest(input_dir: Path) -> dict[str, dict[str, str]]:
    manifest_path = input_dir / "manifest.tsv"
    if not manifest_path.exists():
        return {}
    with manifest_path.open("r", encoding="utf-8-sig", newline="") as file:
        return {
            row["filename"]: row
            for row in csv.DictReader(file, delimiter="\t")
            if row.get("filename")
        }


def load_excel_context(
    excel_path: Path,
    sheet_name: str,
) -> tuple[list[str], dict[int, dict[str, Any]]]:
    """Read source headers and row metadata without extracting embedded images."""
    if not excel_path.exists():
        return [], {}
    workbook = load_workbook(excel_path, read_only=True, data_only=False)
    if sheet_name not in workbook.sheetnames:
        return [], {}
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    metadata = {
        row_number: dict(zip(headers, row_values))
        for row_number, row_values in enumerate(rows, start=2)
    }
    return headers, metadata


def metadata_for_candidate(
    candidate: Path,
    manifest: dict[str, dict[str, str]],
    excel_metadata: dict[int, dict[str, Any]],
) -> dict[str, Any]:
    row: dict[str, Any] = {}
    manifest_row = manifest.get(candidate.name, {})
    row_number = manifest_row.get("row")
    if row_number and row_number.isdigit():
        row["__source_row"] = int(row_number)
        row.update(excel_metadata.get(int(row_number), {}))
    if manifest_row:
        row.setdefault("ASIN", manifest_row.get("asin"))
        row.setdefault("图片链接", manifest_row.get("original_url") or manifest_row.get("used_url_size"))
    if "类别" not in row and row.get("类目路径"):
        row["类别"] = row.get("类目路径")
    row.setdefault("ASIN", extract_asin_from_stem(candidate.stem))
    row.setdefault("图片链接", candidate.name)
    return row


def extract_price_trend_images(
    excel_path: Path,
    sheet_name: str,
    cache_dir: Path,
    source_rows: set[int],
) -> dict[int, Path]:
    if not excel_path.exists() or not source_rows:
        return {}
    cache_dir.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(excel_path, read_only=False, data_only=False)
    if sheet_name not in workbook.sheetnames:
        return {}
    worksheet = workbook[sheet_name]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[1]]
    if "价格趋势图" not in headers:
        return {}

    # 常规 Excel Drawing 图片可直接通过 openpyxl 锚点定位，优先于底层 XML 解析。
    trend_col = headers.index("价格趋势图") + 1
    anchored_trends: dict[int, Path] = {}
    for image in worksheet._images:
        marker = getattr(getattr(image, "anchor", None), "_from", None)
        if marker is None or marker.col + 1 != trend_col or marker.row + 1 not in source_rows:
            continue
        with PILImage.open(BytesIO(image._data())) as source:
            image_path = cache_dir / f"trend_row_{marker.row + 1}.png"
            source.convert("RGB").save(image_path, "PNG")
        anchored_trends[marker.row + 1] = image_path
    if anchored_trends:
        return anchored_trends

    with ZipFile(excel_path) as archive:
        dispimg_images = extract_dispimg_price_trend_images(worksheet, headers, archive, cache_dir, source_rows)
        if dispimg_images:
            return dispimg_images

        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        workbook_ns = {
            "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
        }
        workbook_rels = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels_root.findall("rel:Relationship", workbook_ns)
        }
        sheet_target = ""
        for sheet in workbook_root.findall(".//main:sheet", workbook_ns):
            if sheet.attrib.get("name") == sheet_name:
                rel_id = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
                sheet_target = workbook_rels.get(rel_id, "")
                break
        if not sheet_target:
            return {}
        sheet_path = f"xl/{sheet_target.lstrip('/')}"
        sheet_rels_path = f"{PurePosixPath(sheet_path).parent}/_rels/{PurePosixPath(sheet_path).name}.rels"
        if sheet_rels_path not in archive.namelist():
            return {}
        sheet_rels_root = ET.fromstring(archive.read(sheet_rels_path))
        drawing_target = ""
        for rel in sheet_rels_root.findall("rel:Relationship", workbook_ns):
            if rel.attrib.get("Type", "").endswith("/drawing"):
                drawing_target = rel.attrib["Target"]
                break
        if not drawing_target:
            return {}
        drawing_path = normalize_zip_path(str(PurePosixPath(sheet_path).parent / drawing_target))
        drawing_rels_path = f"{PurePosixPath(drawing_path).parent}/_rels/{PurePosixPath(drawing_path).name}.rels"
        if drawing_path not in archive.namelist() or drawing_rels_path not in archive.namelist():
            return {}

        drawing_rels_root = ET.fromstring(archive.read(drawing_rels_path))
        drawing_rels = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in drawing_rels_root.findall("rel:Relationship", workbook_ns)
        }
        drawing_root = ET.fromstring(archive.read(drawing_path))
        drawing_ns = {
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        }
        drawing_ns["xdr"] = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
        trend_col = headers.index("价格趋势图")
        trend_images: dict[int, Path] = {}
        anchors = drawing_root.findall("xdr:twoCellAnchor", drawing_ns) + drawing_root.findall("xdr:oneCellAnchor", drawing_ns)
        for anchor in anchors:
            marker = anchor.find("xdr:from", drawing_ns)
            blip = anchor.find(".//a:blip", drawing_ns)
            if marker is None or blip is None:
                continue
            col_node = marker.find("xdr:col", drawing_ns)
            row_node = marker.find("xdr:row", drawing_ns)
            if col_node is None or row_node is None:
                continue
            source_row = int(row_node.text) + 1
            if int(col_node.text) != trend_col or source_row not in source_rows:
                continue
            rel_id = blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
            target = drawing_rels.get(rel_id or "")
            if not target:
                continue
            media_path = normalize_zip_path(str(PurePosixPath(drawing_path).parent / target))
            if media_path not in archive.namelist():
                continue
            with PILImage.open(BytesIO(archive.read(media_path))) as source:
                image_path = cache_dir / f"trend_row_{source_row}.png"
                source.convert("RGB").save(image_path, "PNG")
            trend_images[source_row] = image_path
    return trend_images


def extract_dispimg_price_trend_images(
    worksheet: Any,
    headers: list[str],
    archive: ZipFile,
    cache_dir: Path,
    source_rows: set[int],
) -> dict[int, Path]:
    if "xl/cellimages.xml" not in archive.namelist() or "xl/_rels/cellimages.xml.rels" not in archive.namelist():
        return {}

    if "价格趋势图" not in headers:
        return {}
    trend_col = headers.index("价格趋势图") + 1

    package_ns = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}
    rels_root = ET.fromstring(archive.read("xl/_rels/cellimages.xml.rels"))
    rels = {
        rel.attrib["Id"]: normalize_zip_path(f"xl/{rel.attrib['Target']}")
        for rel in rels_root.findall("rel:Relationship", package_ns)
    }

    cell_ns = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    }
    cellimages_root = ET.fromstring(archive.read("xl/cellimages.xml"))
    image_id_to_media: dict[str, str] = {}
    for pic in cellimages_root.findall(".//xdr:pic", cell_ns):
        name_node = pic.find(".//xdr:cNvPr", cell_ns)
        blip = pic.find(".//a:blip", cell_ns)
        if name_node is None or blip is None:
            continue
        image_id = name_node.attrib.get("name")
        rel_id = blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
        media_path = rels.get(rel_id or "")
        if image_id and media_path:
            image_id_to_media[image_id] = media_path

    trend_images: dict[int, Path] = {}
    pattern = re.compile(r'DISPIMG\("([^"]+)"')
    for row in sorted(source_rows):
        value = str(worksheet.cell(row=row, column=trend_col).value or "")
        match = pattern.search(value)
        if not match:
            continue
        media_path = image_id_to_media.get(match.group(1))
        if not media_path or media_path not in archive.namelist():
            continue
        with PILImage.open(BytesIO(archive.read(media_path))) as source:
            image_path = cache_dir / f"trend_row_{row}.png"
            source.convert("RGB").save(image_path, "PNG")
        trend_images[row] = image_path
    return trend_images


def default_compare_prompt_template() -> str:
    return """Compare the target product image with the competitor product image.
Return only valid JSON with these keys:
{
  "color相似度": 0,
  "style相似度": 0,
  "elements相似度": 0,
  "format相似度": 0,
  "相似理由": ""
}
Similarity threshold: {threshold}%.
"""


def load_prompt_template(prompt_path: Path) -> str:
    if prompt_path.exists():
        return prompt_path.read_text(encoding="utf-8")
    return default_compare_prompt_template()


def build_compare_prompt(threshold: float, prompt_template: str) -> str:
    return prompt_template.replace("{threshold}", f"{threshold:g}")


def normalize_similarity_payload(
    data: dict[str, Any],
    threshold: float,
    weights: SimilarityWeights,
) -> dict[str, str]:
    color = parse_percent(data.get("color相似度", data.get("颜色相似度", data.get("color_similarity"))))
    style = parse_percent(data.get("style相似度", data.get("风格相似度", data.get("style_similarity"))))
    elements = parse_percent(data.get("elements相似度", data.get("元素相似度", data.get("elements_similarity"))))
    fmt = parse_percent(data.get("format相似度", data.get("版式相似度", data.get("format_similarity"))))
    # 综合相似度始终由后端按本次任务权重计算，不依赖模型自行加权。
    overall = weights.score(color=color, style=style, elements=elements, format=fmt)
    return {
        "综合相似度": f"{overall:.1f}%",
        "相似理由": str(data.get("相似理由", data.get("reason", ""))).strip(),
        "color相似度": f"{round(color)}%",
        "style相似度": f"{round(style)}%",
        "elements相似度": f"{round(elements)}%",
        "format相似度": f"{round(fmt)}%",
        "是否相似": "是" if overall >= threshold else "否",
    }


def compare_pair(
    client: Any,
    model: str,
    target_url: str,
    candidate_url: str,
    image_detail: str,
    threshold: float,
    prompt_template: str,
    weights: SimilarityWeights,
) -> dict[str, str]:
    response = client.chat.completions.create(
        model=model,
        response_format={"type": "json_object"},
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": build_compare_prompt(threshold, prompt_template)},
                    {"type": "image_url", "image_url": {"url": target_url, "detail": image_detail}},
                    {"type": "image_url", "image_url": {"url": candidate_url, "detail": image_detail}},
                ],
            }
        ],
    )
    return normalize_similarity_payload(extract_json(response.choices[0].message.content or ""), threshold, weights)


def percent_sort_value(value: Any) -> float:
    try:
        return parse_percent(value)
    except Exception:
        return -1.0


def add_image(ws: Any, image_path: Path, cell: str, size: int = 220) -> None:
    """Embed the downloaded original, while keeping its worksheet display compact."""
    image = ExcelImage(str(image_path))
    image.width = size
    image.height = size
    ws.add_image(image, cell)


def add_resized_image(ws: Any, image_path: Path, cell: str, width: int, height: int) -> None:
    image = ExcelImage(str(image_path))
    image.width = width
    image.height = height
    ws.add_image(image, cell)


def process_candidate(
    candidate: Path,
    client: Any,
    model: str,
    target_url: str,
    image_detail: str,
    threshold: float,
    prompt_template: str,
    weights: SimilarityWeights,
) -> dict[str, Any]:
    started = time.perf_counter()
    try:
        similarities = compare_pair(
            client,
            model,
            target_url,
            image_to_data_url(candidate),
            image_detail,
            threshold,
            prompt_template,
            weights,
        )
        error = ""
    except Exception as exc:
        similarities = {
            "综合相似度": "",
            "相似理由": "",
            "color相似度": "",
            "style相似度": "",
            "elements相似度": "",
            "format相似度": "",
            "是否相似": "",
        }
        error = str(exc)
    return {
        "path": candidate,
        "similarities": similarities,
        "error": error,
        "elapsed": time.perf_counter() - started,
    }


def main() -> int:
    from dotenv import load_dotenv

    args = parse_args()
    load_dotenv(PROJECT_ROOT / ".env")
    if args.max_candidates < 1:
        raise ValueError("--max-candidates must be at least 1.")
    if args.max_results < 1:
        raise ValueError("--max-results must be at least 1.")
    try:
        weights = SimilarityWeights.from_json(args.weights) if args.weights else SimilarityWeights()
    except (TypeError, ValueError, json.JSONDecodeError) as exc:
        raise ValueError("--weights 必须是总和为 100 的四维 JSON 对象。") from exc

    input_dir = Path(args.input_dir)
    output_path = Path(args.output)
    run_log_path = Path(args.run_log)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    run_log_path.parent.mkdir(parents=True, exist_ok=True)
    run_log_path.write_text("", encoding="utf-8")

    target_image = find_target_image(input_dir)
    all_candidates = list_candidate_images(input_dir, target_image)
    if not all_candidates:
        raise RuntimeError(f"No candidate images found in {input_dir}")
    # 在调用模型前完成抽样，目标图片不计入候选数量。
    candidates = sample_candidates(all_candidates, args.max_candidates)
    manifest = load_manifest(input_dir)
    metadata_excel_path = Path(args.metadata_excel)
    _, excel_metadata = load_excel_context(
        metadata_excel_path,
        args.metadata_sheet,
    )

    settings = ModelSettings.from_environment()
    model = args.model or settings.similarity_model
    image_detail = args.image_detail or settings.image_detail
    client = create_client(settings, timeout=args.request_timeout, max_retries=args.max_retries)
    prompt_template = load_prompt_template(Path(args.prompt))

    target_url = image_to_data_url(target_image)
    run_started = time.perf_counter()
    sampling_message = f"候选图片: 总数 {len(all_candidates)}，本次识别 {len(candidates)}"
    if len(candidates) < len(all_candidates):
        sampling_message += "（随机抽样）"
    print(sampling_message)
    with run_log_path.open("a", encoding="utf-8") as log:
        log.write(sampling_message + "\n")
        log.write(f"相似度权重: {json.dumps(weights.as_dict(), ensure_ascii=False)}\n")

    results: dict[Path, dict[str, Any]] = {}
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {
            executor.submit(
                process_candidate,
                candidate,
                client,
                model,
                target_url,
                image_detail,
                args.similarity_threshold,
                prompt_template,
                weights,
            ): candidate
            for candidate in candidates
        }
        done = 0
        for future in as_completed(futures):
            candidate = futures[future]
            try:
                result = future.result()
            except Exception as exc:
                # 单个并发任务异常时保留其失败记录，不能中断整批候选的排序。
                result = {
                    "path": candidate,
                    "similarities": {},
                    "error": f"工作线程异常：{exc}",
                    "elapsed": 0.0,
                }
            results[candidate] = result
            done += 1
            with run_log_path.open("a", encoding="utf-8") as log:
                log.write(
                    f"{candidate.name}\t{json.dumps(result['similarities'], ensure_ascii=False)}"
                    f"\t运行时间: {format_duration(result['elapsed'])}"
                    + (f"\t错误: {result['error']}" if result["error"] else "")
                    + "\n"
                )
            print(f"[{done}/{len(candidates)}] {candidate.name} {json.dumps(result['similarities'], ensure_ascii=False)}")

    successful_candidates = [
        candidate
        for candidate in candidates
        if not results[candidate]["error"] and results[candidate]["similarities"].get("综合相似度")
    ]
    if not successful_candidates:
        raise RuntimeError("相似度排序没有成功结果，拒绝生成后续中间数据。")
    sorted_candidates = sorted(
        successful_candidates,
        key=lambda candidate: percent_sort_value(results[candidate]["similarities"].get("综合相似度")),
        reverse=True,
    )[: args.max_results]

    # 中间 JSON 只保存排序和回填所需数据，高清图和趋势图留到最终 Excel 再写入。
    ranked_items: list[dict[str, Any]] = []
    for candidate in sorted_candidates:
        result = results[candidate]
        similarities = result["similarities"]
        source_row = metadata_for_candidate(candidate, manifest, excel_metadata)
        ranked_items.append({
            "source_row": source_row.get("__source_row"),
            "asin": source_row.get("ASIN") or extract_asin_from_stem(candidate.stem),
            "image_url": source_row.get("图片链接", ""),
            "image_name": candidate.name,
            "overall_similarity": similarities["综合相似度"],
            "similarities": similarities,
            "error": result["error"],
        })
    intermediate_result = {
        "schema_version": 1,
        "target_image": target_image.name,
        "candidate_count": len(candidates),
        "result_count": len(ranked_items),
        "threshold": args.similarity_threshold,
        "weights": weights.as_dict(),
        "items": ranked_items,
    }
    output_path.write_text(json.dumps(intermediate_result, ensure_ascii=False, indent=2), encoding="utf-8")
    with run_log_path.open("a", encoding="utf-8") as log:
        log.write(f"本次运行总时间: {format_duration(time.perf_counter() - run_started)}\n")
    print(f"Target image: {target_image}")
    print(f"Output: {output_path}")
    print(f"Run log: {run_log_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
