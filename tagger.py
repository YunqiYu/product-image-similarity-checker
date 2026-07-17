from __future__ import annotations

import argparse
import html
import json
import os
import re
import shutil
import sys
import time
import urllib.request
from zipfile import ZipFile
from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any

DEFAULT_EXCEL = Path(__file__).parent / "inputs" / "关键词开发结果_blue hydrangea_['napkins'].xlsx"
DEFAULT_SHEET = "最终结果"
DEFAULT_URL_COLUMN = "图片链接"
DEFAULT_CSE_PROMPT = Path(__file__).parent / "config" / "商品标签分析任务.docx"
DEFAULT_FORMAT_PROMPT = Path(__file__).parent / "config" / "商品标签识别规范_format_WordArt_最新版.md"
DEFAULT_OUTPUT = Path(__file__).parent / "outputs" / "tagged_results.xlsx"
DEFAULT_FILTER_COLUMN = "是否符合"
DEFAULT_FILTER_VALUE = "符合"
DEFAULT_RUN_LOG = Path(__file__).parent / "outputs" / "run.log"
DEFAULT_WORKERS = 3
DEFAULT_SAVE_EVERY = 20
DEFAULT_TAG_MODE = "split"
DEFAULT_REQUEST_TIMEOUT = 60.0
DEFAULT_MAX_RETRIES = 1

RESULT_COLUMNS = [
    "color",
    "style",
    "elements",
    "format",
    "wordArt",
]

CSE_KEYS = ["color", "style", "elements"]
FORMAT_KEYS = ["format", "wordArt"]
REQUIRED_KEYS = CSE_KEYS + FORMAT_KEYS
OUTPUT_BASE_COLUMNS = ["1级分类", "主题标签", "ASIN", "图片链接", "图片"]
OUTPUT_COLUMNS = OUTPUT_BASE_COLUMNS + RESULT_COLUMNS
SOURCE_COLUMN_ALIASES = {
    "1级分类": ["1级分类", "一级分类"],
    "主题标签": ["主题标签"],
    "ASIN": ["ASIN", "asin"],
    "图片链接": ["图片链接", "image", "图片URL", "图片url"],
}


def upgrade_amazon_image_url(url: str) -> str:
    """Use a higher-resolution Amazon CDN variant when the URL shape allows it."""
    match = re.match(r"^(https://[^?]+?/images/I/[^._]+)(?:\._[^.]+_)?\.jpg$", url)
    if match:
        return match.group(1) + "._AC_SL1500_.jpg"
    return url


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Read image URLs from Excel and generate 5D product image tags."
    )
    parser.add_argument("--excel", default=DEFAULT_EXCEL, help="Excel file path.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Worksheet name.")
    parser.add_argument("--url-column", default=DEFAULT_URL_COLUMN, help="Image URL column title.")
    parser.add_argument("--cse-prompt", default=str(DEFAULT_CSE_PROMPT), help="Color/style/elements prompt path.")
    parser.add_argument("--format-prompt", default=str(DEFAULT_FORMAT_PROMPT), help="Format/WordArt prompt path.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output Excel path.")
    parser.add_argument("--model", default=None, help="Model name. Defaults to OPENAI_MODEL from .env.")
    parser.add_argument("--limit", type=int, default=None, help="Maximum rows to process.")
    parser.add_argument("--start-row", type=int, default=2, help="First worksheet row to process.")
    parser.add_argument("--sleep", type=float, default=0.0, help="Seconds to sleep between requests.")
    parser.add_argument("--overwrite", action="store_true", help="Reprocess rows that already have 五维JSON.")
    parser.add_argument("--filter-column", default=DEFAULT_FILTER_COLUMN, help="Only process rows where this column matches --filter-value.")
    parser.add_argument("--filter-value", default=DEFAULT_FILTER_VALUE, help="Only process rows matching this value.")
    parser.add_argument("--no-filter", action="store_true", help="Process all rows instead of filtering.")
    parser.add_argument("--run-log", default=str(DEFAULT_RUN_LOG), help="Write per-row JSON and timing log to this path.")
    parser.add_argument("--workers", type=int, default=DEFAULT_WORKERS, help="Number of image rows to process concurrently.")
    parser.add_argument("--request-timeout", type=float, default=DEFAULT_REQUEST_TIMEOUT, help="OpenAI-compatible API request timeout in seconds.")
    parser.add_argument("--max-retries", type=int, default=DEFAULT_MAX_RETRIES, help="OpenAI-compatible API SDK retry count.")
    parser.add_argument(
        "--tag-mode",
        choices=["single", "split"],
        default=DEFAULT_TAG_MODE,
        help="single is faster with one model call per image; split uses two calls for CSE and format/WordArt.",
    )
    parser.add_argument(
        "--image-detail",
        default=None,
        help="Image detail passed to the vision API. Defaults to OPENAI_IMAGE_MODE or auto.",
    )
    parser.add_argument(
        "--save-every",
        type=int,
        default=DEFAULT_SAVE_EVERY,
        help="Save output workbook after this many completed rows. Use 1 for safest but slower saving.",
    )
    return parser.parse_args()


def read_prompt(path: str) -> str:
    prompt_path = Path(path)
    if not prompt_path.exists():
        raise FileNotFoundError(f"Prompt file not found: {prompt_path}")
    if prompt_path.suffix.lower() == ".docx":
        return read_docx_text(prompt_path)
    return prompt_path.read_text(encoding="utf-8")


def read_docx_text(path: Path) -> str:
    with ZipFile(path) as docx:
        xml = docx.read("word/document.xml").decode("utf-8")
    paragraphs = re.findall(r"<w:p[\s\S]*?</w:p>", xml)
    lines = []
    for paragraph in paragraphs:
        text_nodes = re.findall(r"<w:t[^>]*>(.*?)</w:t>", paragraph)
        if text_nodes:
            lines.append("".join(html.unescape(text) for text in text_nodes))
    return "\n".join(lines)


def normalize_header(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def find_column_indexes(ws: Any) -> dict[str, int]:
    headers = {}
    for cell in ws[1]:
        title = normalize_header(cell.value)
        if title:
            headers[title] = cell.column
    return headers


def get_source_value(source_ws: Any, row: int, headers: dict[str, int], output_title: str) -> Any:
    for source_title in SOURCE_COLUMN_ALIASES.get(output_title, [output_title]):
        source_col = headers.get(source_title)
        if source_col:
            return source_ws.cell(row=row, column=source_col).value
    return ""


def row_matches_filter(ws: Any, row: int, headers: dict[str, int], args: argparse.Namespace) -> bool:
    if args.no_filter:
        return True
    if not args.filter_column:
        return True
    filter_col = headers.get(args.filter_column)
    if not filter_col:
        print(f"Filter column not found: {args.filter_column}", file=sys.stderr)
        return False
    actual = normalize_header(ws.cell(row=row, column=filter_col).value)
    return actual == args.filter_value


def extract_json(text: str) -> dict[str, Any]:
    stripped = text.strip()
    if stripped.startswith("```"):
        stripped = re.sub(r"^```(?:json)?\s*", "", stripped)
        stripped = re.sub(r"\s*```$", "", stripped)

    try:
        data = json.loads(stripped)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", stripped, flags=re.S)
        if not match:
            raise
        data = json.loads(match.group(0))

    if not isinstance(data, dict):
        raise ValueError("Model output is not a JSON object.")
    return data


def normalize_cse(data: dict[str, Any]) -> dict[str, Any]:
    normalized = OrderedDict()
    for key in CSE_KEYS:
        normalized[key] = data.get(key, [] if key in {"style", "elements"} else "")

    if not isinstance(normalized["style"], list):
        normalized["style"] = [str(normalized["style"])] if normalized["style"] else []
    if not isinstance(normalized["elements"], list):
        normalized["elements"] = [str(normalized["elements"])] if normalized["elements"] else []

    normalized["color"] = str(normalized["color"]).strip()
    normalized["style"] = [str(item).strip() for item in normalized["style"] if str(item).strip()]
    normalized["elements"] = [str(item).strip() for item in normalized["elements"] if str(item).strip()]
    return dict(normalized)


def normalize_format_wordart(data: dict[str, Any]) -> dict[str, Any]:
    word_art = data.get("wordArt", data.get("WordArt", ""))
    if isinstance(word_art, list):
        word_art = " / ".join(str(item).strip() for item in word_art if str(item).strip())

    normalized = OrderedDict()
    normalized["format"] = str(data.get("format", "")).strip()
    normalized["wordArt"] = str(word_art).strip()
    return dict(normalized)


def empty_result() -> dict[str, Any]:
    return {"color": "", "style": [], "elements": [], "format": "", "wordArt": ""}


def merge_results(cse_result: dict[str, Any], format_result: dict[str, Any]) -> dict[str, Any]:
    normalized = OrderedDict()
    normalized["color"] = cse_result.get("color", "")
    normalized["style"] = cse_result.get("style", [])
    normalized["elements"] = cse_result.get("elements", [])
    normalized["format"] = format_result.get("format", "")
    normalized["wordArt"] = format_result.get("wordArt", "")
    return dict(normalized)


def format_duration(seconds: float) -> str:
    total_seconds = int(round(seconds))
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def append_run_log(log_path: Path, message: str) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("a", encoding="utf-8") as file:
        file.write(message + "\n")


def log_row_result(log_path: Path, source_row: int, result: dict[str, Any], elapsed_seconds: float) -> None:
    payload = json.dumps(result, ensure_ascii=False)
    append_run_log(log_path, f"row{{{source_row}}}\t{payload}\t运行时间: {format_duration(elapsed_seconds)}")


def log_total_runtime(log_path: Path, elapsed_seconds: float) -> None:
    append_run_log(log_path, f"本次运行总时间: {format_duration(elapsed_seconds)}")


def build_cse_user_text(prompt: str) -> str:
    return (
        "请严格按照以下商品标签分析任务识别图片，只输出 color、style、elements 三个字段的合法 JSON。\n"
        "不要输出解释、Markdown、代码块或额外字段。\n\n"
        f"{prompt}\n\n"
        "再次确认：最终只输出字段 color, style, elements。"
    )


def build_format_user_text(prompt: str) -> str:
    return (
        "请严格按照以下商品标签识别规范分析图片，只输出 format 和 WordArt 两个字段的合法 JSON。\n"
        "不要输出解释、Markdown、代码块或额外字段。\n\n"
        f"{prompt}\n\n"
        "再次确认：最终只输出字段 format, WordArt。"
    )


def tag_image_with_text(
    client: Any,
    model: str,
    user_text: str,
    image_url: str,
    image_detail: str,
) -> dict[str, Any]:
    response = client.chat.completions.create(
        model=model,
        response_format={"type": "json_object"},
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": user_text},
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": image_url,
                            "detail": image_detail,
                        },
                    },
                ],
            }
        ],
    )
    return extract_json(response.choices[0].message.content or "")


def tag_image_split(
    client: Any,
    model: str,
    cse_prompt: str,
    format_prompt: str,
    image_url: str,
    image_detail: str,
) -> dict[str, Any]:
    cse_result = normalize_cse(
        tag_image_with_text(client, model, build_cse_user_text(cse_prompt), image_url, image_detail)
    )
    format_result = normalize_format_wordart(
        tag_image_with_text(client, model, build_format_user_text(format_prompt), image_url, image_detail)
    )
    return merge_results(cse_result, format_result)


def build_user_text(cse_prompt: str, format_prompt: str) -> str:
    return (
        "请严格按照以下商品图片五维标签识别规范分析图片。\n"
        "只输出一个合法 JSON 对象，不要输出解释、Markdown 或额外字段。\n\n"
        "【Color / Style / Elements 识别规范】\n"
        f"{cse_prompt}\n\n"
        "【Format / WordArt 识别规范】\n"
        f"{format_prompt}\n\n"
        "再次确认：最终输出字段顺序必须是 color, style, elements, format, wordArt。"
    )


def tag_image(
    client: Any,
    model: str,
    cse_prompt: str,
    format_prompt: str,
    image_url: str,
    image_detail: str,
) -> dict[str, Any]:
    data = tag_image_with_text(
        client,
        model,
        build_user_text(cse_prompt, format_prompt),
        image_url,
        image_detail,
    )
    normalized = OrderedDict()
    for key in REQUIRED_KEYS:
        normalized[key] = data.get(key, [] if key in {"style", "elements", "wordArt"} else "")

    if not isinstance(normalized["style"], list):
        normalized["style"] = [str(normalized["style"])] if normalized["style"] else []
    if not isinstance(normalized["elements"], list):
        normalized["elements"] = [str(normalized["elements"])] if normalized["elements"] else []
    if not isinstance(normalized["wordArt"], list):
        normalized["wordArt"] = [str(normalized["wordArt"])] if normalized["wordArt"] else []

    normalized["color"] = str(normalized["color"]).strip()
    normalized["format"] = str(normalized["format"]).strip()
    return dict(normalized)


def format_list(value: Any) -> str:
    if isinstance(value, list):
        return "、".join(str(item).strip() for item in value if str(item).strip())
    return str(value).strip() if value is not None else ""


def create_output_workbook() -> Any:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "最终结果"
    ws.append(OUTPUT_COLUMNS)
    widths = {
        "A": 14,
        "B": 36,
        "C": 16,
        "D": 58,
        "E": 22,
        "F": 12,
        "G": 18,
        "H": 26,
        "I": 28,
        "J": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    return wb


def download_preview_image(image_url: str, image_dir: Path, row: int) -> Path:
    from PIL import Image as PILImage

    image_url = upgrade_amazon_image_url(image_url)
    image_dir.mkdir(parents=True, exist_ok=True)
    raw_path = image_dir / f"row_{row}_raw"
    preview_path = image_dir / f"row_{row}.jpg"

    request = urllib.request.Request(
        image_url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        },
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        raw_path.write_bytes(response.read())

    with PILImage.open(raw_path) as image:
        image = image.convert("RGB")
        image.thumbnail((960, 960))
        image.save(preview_path, format="JPEG", quality=95, optimize=True)

    raw_path.unlink(missing_ok=True)
    return preview_path


def insert_preview_image(out_ws: Any, image_path: Path, output_row: int) -> None:
    from openpyxl.drawing.image import Image as ExcelImage

    image = ExcelImage(str(image_path))
    image.width = 220
    image.height = 220
    out_ws.row_dimensions[output_row].height = 170
    out_ws.add_image(image, f"E{output_row}")


def append_output_row(
    out_ws: Any,
    source_ws: Any,
    row: int,
    headers: dict[str, int],
    result: dict[str, Any],
    image_path: Path | None = None,
) -> None:
    values = []
    for title in OUTPUT_BASE_COLUMNS:
        if title == "图片":
            values.append("")
            continue
        values.append(get_source_value(source_ws, row, headers, title))

    values.extend(
        [
            result["color"],
            format_list(result["style"]),
            format_list(result["elements"]),
            result["format"],
            format_list(result["wordArt"]),
        ]
    )
    out_ws.append(values)
    if image_path is not None:
        insert_preview_image(out_ws, image_path, out_ws.max_row)


def write_output_row(
    out_ws: Any,
    output_row: int,
    source_ws: Any,
    source_row: int,
    headers: dict[str, int],
    result: dict[str, Any],
    image_path: Path | None = None,
) -> None:
    values = []
    for title in OUTPUT_BASE_COLUMNS:
        if title == "图片":
            values.append("")
            continue
        values.append(get_source_value(source_ws, source_row, headers, title))

    values.extend(
        [
            result["color"],
            format_list(result["style"]),
            format_list(result["elements"]),
            result["format"],
            format_list(result["wordArt"]),
        ]
    )

    for col_index, value in enumerate(values, start=1):
        out_ws.cell(row=output_row, column=col_index).value = value
    if image_path is not None:
        insert_preview_image(out_ws, image_path, output_row)


def collect_tasks(ws: Any, headers: dict[str, int], args: argparse.Namespace) -> list[dict[str, Any]]:
    tasks = []
    url_col = headers[args.url_column]
    for row in range(max(args.start_row, 2), ws.max_row + 1):
        if args.limit is not None and len(tasks) >= args.limit:
            break

        image_url = normalize_header(ws.cell(row=row, column=url_col).value)
        if not image_url:
            continue
        if not row_matches_filter(ws, row, headers, args):
            continue

        tasks.append({"index": len(tasks) + 1, "row": row, "image_url": image_url})
    return tasks


def process_task(
    task: dict[str, Any],
    client: Any,
    model: str,
    cse_prompt: str,
    format_prompt: str,
    image_dir: Path,
    tag_mode: str,
    image_detail: str,
) -> dict[str, Any]:
    row_started_at = time.perf_counter()
    image_path = None
    error = None
    recognition_url = upgrade_amazon_image_url(task["image_url"])
    try:
        image_path = download_preview_image(recognition_url, image_dir, task["row"])
    except Exception as exc:
        error = f"image download error {exc}"

    result = empty_result()
    try:
        if tag_mode == "split":
            result = tag_image_split(
                client,
                model,
                cse_prompt,
                format_prompt,
                recognition_url,
                image_detail,
            )
        else:
            result = tag_image(
                client,
                model,
                cse_prompt,
                format_prompt,
                recognition_url,
                image_detail,
            )
    except Exception as exc:
        error = str(exc) if error is None else f"{error}; tag error {exc}"

    return {
        **task,
        "result": result,
        "image_path": image_path,
        "error": error,
        "elapsed": time.perf_counter() - row_started_at,
    }


def main() -> int:
    args = parse_args()

    try:
        from dotenv import load_dotenv
        from openai import OpenAI
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        print(
            "Missing dependency. Run: pip install -r requirements.txt\n"
            f"Detail: {exc}",
            file=sys.stderr,
        )
        return 2

    load_dotenv()

    api_key = os.getenv("OPENAI_API_KEY") or os.getenv("TEXT_API_KEY")
    if not api_key:
        print("Missing OPENAI_API_KEY or TEXT_API_KEY. Create .env and fill it in.", file=sys.stderr)
        return 2

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"Excel file not found: {excel_path}", file=sys.stderr)
        return 2

    cse_prompt = read_prompt(args.cse_prompt)
    format_prompt = read_prompt(args.format_prompt)
    model = args.model or os.getenv("OPENAI_MODEL") or "gpt-5.5"
    image_detail = args.image_detail or os.getenv("OPENAI_IMAGE_MODE") or "auto"
    base_url = os.getenv("OPENAI_BASE_URL") or os.getenv("LLM_BASE_URL")
    client = OpenAI(api_key=api_key, base_url=base_url, timeout=args.request_timeout, max_retries=args.max_retries)

    wb = load_workbook(excel_path)
    if args.sheet not in wb.sheetnames:
        print(f"Sheet not found: {args.sheet}. Available sheets: {', '.join(wb.sheetnames)}", file=sys.stderr)
        return 2

    ws = wb[args.sheet]
    headers = find_column_indexes(ws)
    if args.url_column not in headers:
        print(f"Column not found: {args.url_column}", file=sys.stderr)
        return 2

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    run_log_path = Path(args.run_log)
    run_log_path.parent.mkdir(parents=True, exist_ok=True)
    run_log_path.write_text("", encoding="utf-8")

    image_dir = output_path.parent / f".{output_path.stem}_image_cache"
    if image_dir.exists():
        shutil.rmtree(image_dir)
    out_wb = create_output_workbook()
    out_ws = out_wb.active

    tasks = collect_tasks(ws, headers, args)
    if args.workers < 1:
        print("--workers must be >= 1", file=sys.stderr)
        return 2
    if args.save_every < 1:
        print("--save-every must be >= 1", file=sys.stderr)
        return 2

    run_started_at = time.perf_counter()
    print(f"Collected {len(tasks)} rows. workers={args.workers} tag_mode={args.tag_mode} image_detail={image_detail}")

    if args.workers == 1:
        for task in tasks:
            print(f"[{task['index']}] row={task['row']} url={task['image_url']}")
            completed = process_task(
                task,
                client,
                model,
                cse_prompt,
                format_prompt,
                image_dir,
                args.tag_mode,
                image_detail,
            )
            append_output_row(out_ws, ws, completed["row"], headers, completed["result"], completed["image_path"])
            log_row_result(run_log_path, completed["row"], completed["result"], completed["elapsed"])
            if completed["error"]:
                print(f"    error {completed['error']}", file=sys.stderr)
            else:
                print(f"    ok {json.dumps(completed['result'], ensure_ascii=False)}")
            if completed["index"] % args.save_every == 0:
                out_wb.save(output_path)
            if args.sleep:
                time.sleep(args.sleep)
    else:
        with ThreadPoolExecutor(max_workers=args.workers) as executor:
            future_map = {}
            for task in tasks:
                print(f"[{task['index']}] queued row={task['row']} url={task['image_url']}")
                future = executor.submit(
                    process_task,
                    task,
                    client,
                    model,
                    cse_prompt,
                    format_prompt,
                    image_dir,
                    args.tag_mode,
                    image_detail,
                )
                future_map[future] = task
                if args.sleep:
                    time.sleep(args.sleep)

            completed_count = 0
            completed_by_index = {}
            for future in as_completed(future_map):
                task = future_map[future]
                try:
                    completed = future.result()
                except Exception as exc:
                    completed = {
                        **task,
                        "result": empty_result(),
                        "image_path": None,
                        "error": str(exc),
                        "elapsed": 0.0,
                    }

                completed_count += 1
                completed_by_index[completed["index"]] = completed
                log_row_result(run_log_path, completed["row"], completed["result"], completed["elapsed"])
                prefix = f"[{completed_count}/{len(tasks)}] row={completed['row']}"
                if completed["error"]:
                    print(f"{prefix} error {completed['error']}", file=sys.stderr)
                else:
                    print(f"{prefix} ok {json.dumps(completed['result'], ensure_ascii=False)}")
                write_output_row(
                    out_ws,
                    completed["index"] + 1,
                    ws,
                    completed["row"],
                    headers,
                    completed["result"],
                    completed["image_path"],
                )
                if completed_count % args.save_every == 0:
                    out_wb.save(output_path)

    out_wb.save(output_path)
    if image_dir.exists():
        shutil.rmtree(image_dir)
    log_total_runtime(run_log_path, time.perf_counter() - run_started_at)
    print(f"Done. Processed {len(tasks)} rows. Output: {output_path}")
    print(f"Run log: {run_log_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
