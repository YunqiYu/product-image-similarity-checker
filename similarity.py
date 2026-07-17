from __future__ import annotations

import argparse
import base64
import json
import mimetypes
import os
import shutil
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage

from tagger import (
    DEFAULT_CSE_PROMPT,
    DEFAULT_FORMAT_PROMPT,
    download_preview_image,
    empty_result,
    find_column_indexes,
    format_duration,
    format_list,
    normalize_header,
    read_prompt,
    tag_image,
    tag_image_split,
    upgrade_amazon_image_url,
)


DEFAULT_TARGET_IMAGE = Path(__file__).parent / "inputs" / "目标图片.png"
DEFAULT_EXCEL = Path(__file__).parent / "inputs" / "一品红测试.xlsx"
DEFAULT_SHEET = "Sheet1"
DEFAULT_URL_COLUMN = "图片链接"
DEFAULT_OUTPUT = Path(__file__).parent / "outputs" / "一品红测试_相似度结果.xlsx"
DEFAULT_RUN_LOG = Path(__file__).parent / "outputs" / "一品红测试_相似度结果.log"

SIMILARITY_COLUMNS = [
    "元素相似度",
    "风格相似度",
    "颜色相似度",
    "版式相似度",
    "综合相似度",
    "是否相似",
    "相似理由",
    "color相似度",
    "style相似度",
    "elements相似度",
    "format相似度",
    "wordArt相似度",
]

ELEMENT_SYNONYMS = {
    "一品红": "圣诞花",
    "圣诞红": "圣诞花",
    "松枝": "圣诞植物",
    "冬青": "圣诞植物",
    "圣诞叶": "圣诞植物",
    "叶片": "叶子",
    "绿叶": "叶子",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare one target image against Excel images by 5D tags.")
    parser.add_argument("--target-image", default=str(DEFAULT_TARGET_IMAGE))
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL))
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--url-column", default=DEFAULT_URL_COLUMN)
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--run-log", default=str(DEFAULT_RUN_LOG))
    parser.add_argument("--model", default=None)
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--workers", type=int, default=3)
    parser.add_argument("--tag-mode", choices=["single", "split"], default="split")
    parser.add_argument("--image-detail", default=None)
    parser.add_argument("--request-timeout", type=float, default=60.0)
    parser.add_argument("--max-retries", type=int, default=1)
    parser.add_argument("--compare-mode", choices=["direct", "tags"], default="direct")
    parser.add_argument("--similarity-threshold", type=float, default=80.0)
    return parser.parse_args()


def find_existing_path(path_text: str, pattern: str) -> Path:
    path = Path(path_text)
    if path.exists():
        return path
    candidates = sorted((Path(__file__).parent / "inputs").glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    if candidates:
        return candidates[0]
    raise FileNotFoundError(path_text)


def local_image_to_data_url(path: Path) -> str:
    mime_type = mimetypes.guess_type(path.name)[0] or "image/png"
    data = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime_type};base64,{data}"


def split_items(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item).strip().lower() for item in value if str(item).strip()]
    text = str(value or "").strip()
    if not text:
        return []
    for sep in ["、", ",", "，", "/", "|"]:
        text = text.replace(sep, "、")
    return [item.strip().lower() for item in text.split("、") if item.strip()]


def normalize_element(item: str) -> str:
    item = item.strip().lower()
    return ELEMENT_SYNONYMS.get(item, item)


def percent(value: float) -> str:
    value = max(0.0, min(1.0, value))
    return f"{round(value * 100)}%"


def text_similarity(a: Any, b: Any) -> float:
    a_text = str(a or "").strip().lower()
    b_text = str(b or "").strip().lower()
    if not a_text and not b_text:
        return 1.0
    if not a_text or not b_text:
        return 0.0
    if a_text == b_text:
        return 1.0
    return SequenceMatcher(None, a_text, b_text).ratio()


def set_similarity(a: Any, b: Any) -> float:
    a_set = set(split_items(a))
    b_set = set(split_items(b))
    if not a_set and not b_set:
        return 1.0
    if not a_set or not b_set:
        return 0.0
    return len(a_set & b_set) / len(a_set | b_set)


def style_similarity(a: Any, b: Any) -> float:
    a_set = set(split_items(a))
    b_set = set(split_items(b))
    if not a_set and not b_set:
        return 1.0
    if not a_set or not b_set:
        return 0.0
    if a_set <= b_set or b_set <= a_set:
        return 1.0
    return len(a_set & b_set) / len(a_set | b_set)


def element_similarity(a: Any, b: Any) -> float:
    a_set = {normalize_element(item) for item in split_items(a)}
    b_set = {normalize_element(item) for item in split_items(b)}
    if not a_set and not b_set:
        return 1.0
    if not a_set or not b_set:
        return 0.0
    if a_set <= b_set or b_set <= a_set:
        return 1.0
    return len(a_set & b_set) / len(a_set | b_set)


def format_similarity(a: Any, b: Any) -> float:
    a_text = str(a or "").strip()
    b_text = str(b or "").strip()
    if not a_text and not b_text:
        return 1.0
    if a_text == b_text:
        return 1.0
    a_prefix = a_text.split("-", 1)[0]
    b_prefix = b_text.split("-", 1)[0]
    if a_prefix and a_prefix == b_prefix:
        return 0.6
    return 0.0


def calculate_similarities(target: dict[str, Any], candidate: dict[str, Any]) -> dict[str, str]:
    similarities = {
        "color相似度": percent(text_similarity(target.get("color"), candidate.get("color"))),
        "style相似度": percent(style_similarity(target.get("style"), candidate.get("style"))),
        "elements相似度": percent(element_similarity(target.get("elements"), candidate.get("elements"))),
        "format相似度": percent(format_similarity(target.get("format"), candidate.get("format"))),
        "wordArt相似度": percent(text_similarity(target.get("wordArt"), candidate.get("wordArt"))),
    }
    return {
        "元素相似度": similarities["elements相似度"],
        "风格相似度": similarities["style相似度"],
        "颜色相似度": similarities["color相似度"],
        "版式相似度": similarities["format相似度"],
        "综合相似度": "0%",
        "是否相似": "",
        "相似理由": "",
        **similarities,
    }


def parse_percent(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "").strip().replace("%", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def normalize_similarity_payload(data: dict[str, Any], threshold: float) -> dict[str, str]:
    element = parse_percent(data.get("元素相似度", data.get("element_similarity", data.get("elements_similarity"))))
    style = parse_percent(data.get("风格相似度", data.get("style_similarity")))
    color = parse_percent(data.get("颜色相似度", data.get("color_similarity")))
    layout = parse_percent(data.get("版式相似度", data.get("format_similarity", data.get("layout_similarity"))))
    overall = element * 0.4 + style * 0.2 + color * 0.2 + layout * 0.2
    return {
        "元素相似度": f"{round(element)}%",
        "风格相似度": f"{round(style)}%",
        "颜色相似度": f"{round(color)}%",
        "版式相似度": f"{round(layout)}%",
        "综合相似度": f"{overall:.1f}%",
        "是否相似": "是" if overall >= threshold else "否",
        "相似理由": str(data.get("相似理由", data.get("reason", ""))).strip(),
        "color相似度": f"{round(color)}%",
        "style相似度": f"{round(style)}%",
        "elements相似度": f"{round(element)}%",
        "format相似度": f"{round(layout)}%",
        "wordArt相似度": "",
    }


def build_direct_compare_prompt(threshold: float) -> str:
    return f"""你是资深商品图片视觉相似度评估专家。
现在给你两张商品图片：
- 第一张是目标图；
- 第二张是候选图。

请只评估商品正面印刷图案，不要因为商品类型、摆放角度、盘子/纸巾数量、包装组合、拍摄背景、阴影、页面文字而扣分。

请从四个维度分别给出 0-100 的相似度：
1. 元素相似度：权重40%。比较主体元素和主题元素是否一致或语义接近，例如一品红/圣诞花、松枝/冬青/圣诞叶可以视为高度相近。
2. 风格相似度：权重20%。比较插画、矢量、水彩、写实、特效等视觉风格是否接近。
3. 颜色相似度：权重20%。比较主要视觉配色是否接近，忽略白色摄影背景。
4. 版式相似度：权重20%。比较散排、居中、环绕、铺满、文案居中等构图关系是否接近。

综合相似度公式：
综合相似度 = 元素相似度*40% + 风格相似度*20% + 颜色相似度*20% + 版式相似度*20%

如果两张图片本质上是同一设计、同一主图或只是分辨率不同，应给出接近或等于100%的分数。
相似阈值为 {threshold}%。

只输出合法 JSON，不要输出解释性段落或 Markdown。格式如下：
{{
  "元素相似度": 95,
  "风格相似度": 95,
  "颜色相似度": 98,
  "版式相似度": 96,
  "相似理由": "两张图在主体元素、风格、颜色和版式上高度一致"
}}
"""


def direct_compare_images(
    client: Any,
    model: str,
    target_image_url: str,
    candidate_image_url: str,
    image_detail: str,
    threshold: float,
) -> dict[str, str]:
    from tagger import extract_json

    response = client.chat.completions.create(
        model=model,
        response_format={"type": "json_object"},
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": build_direct_compare_prompt(threshold)},
                    {"type": "image_url", "image_url": {"url": target_image_url, "detail": image_detail}},
                    {"type": "image_url", "image_url": {"url": candidate_image_url, "detail": image_detail}},
                ],
            }
        ],
    )
    return normalize_similarity_payload(extract_json(response.choices[0].message.content or ""), threshold)


def add_image(ws: Any, image_path: Path, cell: str, size: int = 220) -> None:
    image = ExcelImage(str(image_path))
    image.width = size
    image.height = size
    ws.add_image(image, cell)


def write_target_block(ws: Any, target_image: Path, target_result: dict[str, Any]) -> None:
    headers = ["目标图片", "target_color", "target_style", "target_elements", "target_format", "target_wordArt"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header
    ws.row_dimensions[2].height = 170
    ws.column_dimensions["A"].width = 32
    add_image(ws, target_image, "A2")
    ws.cell(row=2, column=2).value = target_result["color"]
    ws.cell(row=2, column=3).value = format_list(target_result["style"])
    ws.cell(row=2, column=4).value = format_list(target_result["elements"])
    ws.cell(row=2, column=5).value = target_result["format"]
    ws.cell(row=2, column=6).value = format_list(target_result["wordArt"])


def write_candidate_header(ws: Any, source_headers: list[str], start_row: int) -> list[str]:
    output_headers = source_headers + ["识别color", "识别style", "识别elements", "识别format", "识别wordArt"] + SIMILARITY_COLUMNS
    for col, header in enumerate(output_headers, start=1):
        ws.cell(row=start_row, column=col).value = header
    return output_headers


def process_candidate(
    row_index: int,
    image_url: str,
    client: Any,
    model: str,
    cse_prompt: str,
    format_prompt: str,
    image_detail: str,
    tag_mode: str,
    target_result: dict[str, Any],
    target_image_url: str,
    image_dir: Path,
    compare_mode: str,
    threshold: float,
) -> dict[str, Any]:
    started = time.perf_counter()
    image_path = None
    error = None
    recognition_url = upgrade_amazon_image_url(image_url)
    try:
        image_path = download_preview_image(recognition_url, image_dir, row_index)
    except Exception as exc:
        error = f"image download error {exc}"
    try:
        if compare_mode == "direct":
            result = empty_result()
            similarities = direct_compare_images(client, model, target_image_url, recognition_url, image_detail, threshold)
        else:
            if tag_mode == "split":
                result = tag_image_split(client, model, cse_prompt, format_prompt, recognition_url, image_detail)
            else:
                result = tag_image(client, model, cse_prompt, format_prompt, recognition_url, image_detail)
            similarities = calculate_similarities(target_result, result)
    except Exception as exc:
        result = empty_result()
        similarities = normalize_similarity_payload({}, threshold)
        error = str(exc) if error is None else f"{error}; tag error {exc}"
    return {
        "row_index": row_index,
        "result": result,
        "similarities": similarities,
        "image_path": image_path,
        "error": error,
        "elapsed": time.perf_counter() - started,
    }


def main() -> int:
    from dotenv import load_dotenv
    from openai import OpenAI

    args = parse_args()
    load_dotenv(Path(__file__).parent / ".env")

    api_key = os.getenv("OPENAI_API_KEY") or os.getenv("TEXT_API_KEY")
    if not api_key:
        raise RuntimeError("Missing OPENAI_API_KEY or TEXT_API_KEY")
    base_url = os.getenv("OPENAI_BASE_URL") or os.getenv("LLM_BASE_URL")
    model = args.model or os.getenv("OPENAI_MODEL") or "gpt-5.5"
    image_detail = args.image_detail or os.getenv("OPENAI_IMAGE_MODE") or "high"
    client = OpenAI(api_key=api_key, base_url=base_url, timeout=args.request_timeout, max_retries=args.max_retries)

    target_image = find_existing_path(args.target_image, "*目标图片*.png")
    excel_path = find_existing_path(args.excel, "*一品红测试*.xlsx")
    output_path = Path(args.output)
    run_log_path = Path(args.run_log)
    image_dir = output_path.parent / f".{output_path.stem}_image_cache"
    if image_dir.exists():
        shutil.rmtree(image_dir)
    image_dir.mkdir(parents=True, exist_ok=True)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    run_log_path.parent.mkdir(parents=True, exist_ok=True)
    run_log_path.write_text("", encoding="utf-8")

    cse_prompt = read_prompt(str(DEFAULT_CSE_PROMPT))
    format_prompt = read_prompt(str(DEFAULT_FORMAT_PROMPT))

    run_started = time.perf_counter()
    target_url = local_image_to_data_url(target_image)
    if args.compare_mode == "tags":
        if args.tag_mode == "split":
            target_result = tag_image_split(client, model, cse_prompt, format_prompt, target_url, image_detail)
        else:
            target_result = tag_image(client, model, cse_prompt, format_prompt, target_url, image_detail)
    else:
        target_result = empty_result()

    source_wb = load_workbook(excel_path, data_only=True)
    source_ws = source_wb[args.sheet] if args.sheet in source_wb.sheetnames else source_wb.worksheets[0]
    headers = find_column_indexes(source_ws)
    url_col = headers.get(args.url_column)
    if not url_col:
        raise RuntimeError(f"URL column not found: {args.url_column}")

    source_headers = [source_ws.cell(row=1, column=col).value or "" for col in range(1, source_ws.max_column + 1)]
    tasks = []
    for row in range(2, source_ws.max_row + 1):
        if args.limit is not None and len(tasks) >= args.limit:
            break
        image_url = normalize_header(source_ws.cell(row=row, column=url_col).value)
        if image_url:
            tasks.append((row, image_url))

    results = {}
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {
            executor.submit(
                process_candidate,
                row,
                image_url,
                client,
                model,
                cse_prompt,
                format_prompt,
                image_detail,
                args.tag_mode,
                target_result,
                target_url,
                image_dir,
                args.compare_mode,
                args.similarity_threshold,
            ): (row, image_url)
            for row, image_url in tasks
        }
        done = 0
        for future in as_completed(futures):
            row, _ = futures[future]
            done += 1
            item = future.result()
            results[row] = item
            line = (
                f"row{{{row}}}\t{json.dumps(item['result'], ensure_ascii=False)}\t"
                f"{json.dumps(item['similarities'], ensure_ascii=False)}\t运行时间: {format_duration(item['elapsed'])}"
            )
            if item["error"]:
                line += f"\t错误: {item['error']}"
            with run_log_path.open("a", encoding="utf-8") as log:
                log.write(line + "\n")
            print(f"[{done}/{len(tasks)}] row={row} {json.dumps(item['similarities'], ensure_ascii=False)}")

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "相似度结果"
    write_target_block(out_ws, target_image, target_result)
    header_row = 4
    write_candidate_header(out_ws, source_headers, header_row)
    out_ws.column_dimensions["E"].width = 32

    for row in range(2, source_ws.max_row + 1):
        out_row = row + header_row - 1
        for col in range(1, source_ws.max_column + 1):
            value = source_ws.cell(row=row, column=col).value
            if col == headers.get("图片"):
                value = ""
            out_ws.cell(row=out_row, column=col).value = value
        item = results.get(row)
        if not item:
            continue
        result = item["result"]
        start_col = len(source_headers) + 1
        values = [
            result["color"],
            format_list(result["style"]),
            format_list(result["elements"]),
            result["format"],
            format_list(result["wordArt"]),
        ] + [item["similarities"].get(name, "") for name in SIMILARITY_COLUMNS]
        for offset, value in enumerate(values):
            out_ws.cell(row=out_row, column=start_col + offset).value = value
        if item["image_path"] and headers.get("图片"):
            out_ws.row_dimensions[out_row].height = 170
            add_image(out_ws, item["image_path"], f"{source_ws.cell(1, headers['图片']).column_letter}{out_row}")

    out_wb.save(output_path)
    shutil.rmtree(image_dir, ignore_errors=True)
    with run_log_path.open("a", encoding="utf-8") as log:
        log.write(f"本次运行总时间: {format_duration(time.perf_counter() - run_started)}\n")
    print(f"Output: {output_path}")
    print(f"Run log: {run_log_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
